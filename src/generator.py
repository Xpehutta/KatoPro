"""Генерация листов ежемесячных отчётов Excel."""

from __future__ import annotations

import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from .config_loader import AppConfig, PointConfig
from .holidays import HolidaysProvider
from .utils import (
    MONTH_NAMES_RU,
    days_in_month,
    is_weekend,
    month_dates,
    parse_sheet_name,
    previous_month,
    sheet_name_for,
)

GREEN_FILL = PatternFill(start_color="C5E0B2", end_color="C5E0B2", fill_type="solid")
GRAY_FILL = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATA_START_ROW = 14
INPUT_COLS = (2, 3, 5, 6, 7)  # B C E F G
SELLER_COLS = (10, 11, 12, 13, 14, 15)  # J–O

MONEY_FMT = r'_-* #,##0.00\ _₽_-;\-* #,##0.00\ _₽_-;_-* "-"??\ _₽_-;_-@_-'
DATE_FMT = "mm-dd-yy"
HOLIDAY_LIST_COL = 27  # AA — служебный список праздников для NETWORKDAYS


class ReportGenerator:
    def __init__(self, config: AppConfig, holidays: HolidaysProvider):
        self.config = config
        self.holidays = holidays
        Path(config.output_dir).mkdir(parents=True, exist_ok=True)

    def generate(
        self,
        year: int,
        month: int,
        point_names: Optional[list[str]] = None,
        point_periods: Optional[dict[str, tuple[int, int]]] = None,
    ) -> list[dict]:
        targets = self._resolve_points(point_names)
        results: list[dict] = []
        for point in targets:
            y, m = year, month
            if point_periods:
                for key, value in point_periods.items():
                    if key.casefold() == point.name.casefold():
                        y, m = value
                        break
            try:
                results.append(self._generate_point(point, y, m))
            except Exception as exc:
                logger.exception("Ошибка генерации для {}: {}", point.name, exc)
                results.append(
                    {
                        "name": point.name,
                        "status": "error",
                        "output_file": None,
                        "sheet_name": None,
                        "message": str(exc),
                        "year": y,
                        "month": m,
                    }
                )
        return results

    def _resolve_points(self, point_names: Optional[list[str]]) -> list[PointConfig]:
        if not point_names:
            if not self.config.points:
                raise ValueError("В config.yaml не заданы торговые точки")
            return list(self.config.points)

        resolved: list[PointConfig] = []
        missing: list[str] = []
        for name in point_names:
            point = self.config.point_by_name(name)
            if point is None:
                missing.append(name)
            else:
                resolved.append(point)
        if missing:
            raise ValueError(f"Неизвестные точки: {', '.join(missing)}")
        return resolved

    def _generate_point(self, point: PointConfig, year: int, month: int) -> dict:
        src = self._resolve_source_file(point, year, month)
        out_name = f"МесОтч{year}{month:02d}_{point.name}.xlsx"
        dest = Path(self.config.output_dir) / out_name
        if src.resolve() != dest.resolve():
            shutil.copy2(src, dest)
            logger.info("Скопирован {} → {}", src, dest)
        else:
            logger.info("Обновление существующей книги {}", dest)

        wb = load_workbook(dest)
        new_sheet = sheet_name_for(year, month)

        if new_sheet in wb.sheetnames:
            logger.info("Лист {} уже есть — пересоздаём", new_sheet)
            del wb[new_sheet]

        prev_sheet_name = self._resolve_previous_sheet_name(wb, year, month)
        template_ws = wb[prev_sheet_name]

        ws = wb.copy_worksheet(template_ws)
        ws.title = new_sheet

        all_non_working = set(self.holidays.get_month_holidays(year, month))
        # Для NETWORKDAYS нужны праздники/переносы в будни (сб/вс NETWORKDAYS исключает сам).
        weekday_holidays = sorted(d for d in all_non_working if d.weekday() < 5)

        self._rebuild_sheet(
            ws=ws,
            year=year,
            month=month,
            prev_sheet_name=prev_sheet_name,
            prev_ws=template_ws,
            source_path=src,
            non_working=all_non_working,
            weekday_holidays=weekday_holidays,
        )

        self._protect_sheet(ws)
        wb.save(dest)
        wb.close()
        logger.info("Сохранён {} с листом {}", dest, new_sheet)

        return {
            "name": point.name,
            "status": "ok",
            "output_file": str(dest),
            "sheet_name": new_sheet,
            "year": year,
            "month": month,
            "message": (
                f"Создан лист {new_sheet}; "
                f"праздников для NETWORKDAYS: {len(weekday_holidays)}"
            ),
        }

    def _resolve_source_file(self, point: PointConfig, year: int, month: int) -> Path:
        """
        Выбирает книгу с самой полной историей до целевого месяца:
        конфиг + уже сгенерированные файлы этой точки.
        """
        configured = Path(point.file_path)
        candidates: list[Path] = []
        if configured.exists():
            candidates.append(configured)

        out_dir = Path(self.config.output_dir)
        candidates.extend(sorted(out_dir.glob(f"МесОтч*_{point.name}.xlsx")))

        data_dir = configured.parent if configured.parent.exists() else Path("data")
        if data_dir.exists():
            candidates.extend(sorted(data_dir.glob(f"МесОтч*_{point.name}.xlsx")))

        # unique preserve order
        seen: set[Path] = set()
        unique: list[Path] = []
        for path in candidates:
            resolved = path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            unique.append(path)

        if not unique:
            raise FileNotFoundError(f"Исходный файл не найден: {configured}")

        target = (year, month)
        best_path = unique[0]
        best_score = (-1, -1)  # (year, month) of latest sheet before target

        for path in unique:
            try:
                wb = load_workbook(path, read_only=True)
                latest_before = None
                for name in wb.sheetnames:
                    parsed = parse_sheet_name(name)
                    if parsed and parsed < target:
                        if latest_before is None or parsed > latest_before:
                            latest_before = parsed
                wb.close()
                score = latest_before or (0, 0)
                if score > best_score:
                    best_score = score
                    best_path = path
            except Exception as exc:
                logger.warning("Не удалось прочитать {}: {}", path, exc)

        logger.info(
            "Источник для {} {}/{}: {} (последний лист до цели: {}/{})",
            point.name,
            year,
            month,
            best_path,
            best_score[0],
            best_score[1],
        )
        return best_path

    def _resolve_previous_sheet_name(self, wb, year: int, month: int) -> str:
        """Имя листа предыдущего доступного месяца (не обязательно month-1)."""
        preferred = sheet_name_for(*previous_month(year, month))
        if preferred in wb.sheetnames:
            return preferred

        target = (year, month)
        best_name = None
        best_key = None
        for name in wb.sheetnames:
            parsed = parse_sheet_name(name)
            if parsed and parsed < target:
                if best_key is None or parsed > best_key:
                    best_key = parsed
                    best_name = name

        if best_name:
            logger.warning(
                "Лист {} не найден — используем ближайший {}",
                preferred,
                best_name,
            )
            return best_name

        raise ValueError(
            f"Не найден лист предыдущего месяца до {sheet_name_for(year, month)}"
        )

    def _last_cash_balance_row(self, prev_ws: Worksheet) -> Optional[int]:
        """Строка последней даты в колонке A (зона ежедневных данных)."""
        last_row: Optional[int] = None
        max_row = prev_ws.max_row or DATA_START_ROW
        for row in range(DATA_START_ROW, max_row + 1):
            val = prev_ws.cell(row, 1).value
            if isinstance(val, (datetime, date)):
                last_row = row
            elif last_row is not None and val is None:
                break
        return last_row

    def _opening_balance_for_a3(
        self,
        prev_ws: Worksheet,
        prev_sheet_name: str,
        source_path: Optional[Path] = None,
    ):
        """
        Остаток в кассе за последнюю дату предыдущего месяца → в A3.
        Предпочтительно копируем рассчитанное число; иначе — ссылку на ячейку.
        """
        last_row = self._last_cash_balance_row(prev_ws)
        if last_row is None:
            logger.warning(
                "Не найдена последняя дата на листе {} — fallback на I3",
                prev_sheet_name,
            )
            return f"={prev_sheet_name}!I3"

        cell = prev_ws.cell(last_row, 9)  # колонка I — «Остаток в кассе»
        if isinstance(cell.value, (int, float)):
            logger.info(
                "A3 ← число из {}!I{} = {}",
                prev_sheet_name,
                last_row,
                cell.value,
            )
            return cell.value

        # Попробуем кэш рассчитанных значений Excel (data_only)
        if source_path is not None and source_path.exists():
            try:
                cached_wb = load_workbook(source_path, data_only=True)
                if prev_sheet_name in cached_wb.sheetnames:
                    cached_val = cached_wb[prev_sheet_name].cell(last_row, 9).value
                    cached_wb.close()
                    if isinstance(cached_val, (int, float)):
                        logger.info(
                            "A3 ← кэш Excel {}!I{} = {}",
                            prev_sheet_name,
                            last_row,
                            cached_val,
                        )
                        return cached_val
                else:
                    cached_wb.close()
            except Exception as exc:
                logger.warning("Не удалось прочитать data_only из {}: {}", source_path, exc)

        ref = f"={prev_sheet_name}!I{last_row}"
        logger.info("A3 ← ссылка на остаток за последнюю дату: {}", ref)
        return ref

    def _rebuild_sheet(
        self,
        ws: Worksheet,
        year: int,
        month: int,
        prev_sheet_name: str,
        prev_ws: Worksheet,
        source_path: Optional[Path],
        non_working: set[date],
        weekday_holidays: list[date],
    ) -> None:
        n_days = days_in_month(year, month)
        last_data_row = DATA_START_ROW + n_days - 1
        sum_row = last_data_row + 1
        stats_start = sum_row + 2

        self._clear_data_area(ws, DATA_START_ROW, max(ws.max_row or 60, 60))
        self._clear_holiday_helper(ws)

        opening = self._opening_balance_for_a3(prev_ws, prev_sheet_name, source_path)
        ws["A3"] = opening
        ws["A3"].number_format = MONEY_FMT
        ws["D3"] = f"=D{sum_row}"
        ws["E3"] = f"=E{sum_row}"
        ws["G3"] = f"=G{sum_row}"
        ws["I3"] = f"=I{last_data_row}"
        ws["A7"] = year
        ws["B7"] = MONTH_NAMES_RU[month]
        ws["B10"] = f"={prev_sheet_name}!B10"
        ws["B10"].fill = YELLOW_FILL
        ws["I13"] = "=A3"

        date_to_row: dict[date, int] = {}
        for offset, d in enumerate(month_dates(year, month)):
            row = DATA_START_ROW + offset
            date_to_row[d] = row
            is_off = d in non_working or is_weekend(d)
            self._write_day_row(ws, row, d, is_off)

        holiday_arg = self._build_networkdays_arg(ws, weekday_holidays, date_to_row)
        if holiday_arg:
            ws["B8"] = f"=NETWORKDAYS(A{DATA_START_ROW},A{last_data_row},{holiday_arg})"
        else:
            ws["B8"] = f"=NETWORKDAYS(A{DATA_START_ROW},A{last_data_row})"

        for col in ("B", "C", "D", "E", "F", "G"):
            cell = ws[f"{col}{sum_row}"]
            cell.value = f"=SUM({col}{DATA_START_ROW}:{col}{last_data_row})"
            cell.number_format = "0.00" if col in {"B", "C", "D", "E"} else "General"
            cell.protection = Protection(locked=True)
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=10)

        ws[f"C{stats_start}"] = "Кол-во дней"
        ws[f"D{stats_start}"] = "=B8"
        ws[f"C{stats_start + 1}"] = "ср. выручка"
        ws[f"D{stats_start + 1}"] = f"=D{sum_row}/D{stats_start}"
        ws[f"C{stats_start + 2}"] = "ср. чек"
        ws[f"D{stats_start + 2}"] = f"=B{sum_row}/F{sum_row}"
        for r in range(stats_start, stats_start + 3):
            ws[f"C{r}"].protection = Protection(locked=True)
            ws[f"D{r}"].protection = Protection(locked=True)
            if r > stats_start:
                ws[f"D{r}"].number_format = "0.00"

    def _build_networkdays_arg(
        self,
        ws: Worksheet,
        weekday_holidays: list[date],
        date_to_row: dict[date, int],
    ) -> Optional[str]:
        if not weekday_holidays:
            return None

        rows = [date_to_row[d] for d in weekday_holidays if d in date_to_row]
        if not rows:
            return None

        # Непрерывный диапазон в колонке A — как в исходных файлах
        if rows == list(range(rows[0], rows[-1] + 1)):
            if len(rows) == 1:
                return f"A{rows[0]}"
            return f"A{rows[0]}:A{rows[-1]}"

        # Несмежные праздники — список в колонке AA
        for idx, d in enumerate(weekday_holidays, start=1):
            cell = ws.cell(idx, HOLIDAY_LIST_COL, datetime(d.year, d.month, d.day))
            cell.number_format = DATE_FMT
            cell.protection = Protection(locked=True)
        if len(weekday_holidays) == 1:
            return "AA1"
        return f"AA1:AA{len(weekday_holidays)}"

    def _write_day_row(self, ws: Worksheet, row: int, d: date, is_off: bool) -> None:
        a = ws.cell(row, 1, datetime(d.year, d.month, d.day))
        a.number_format = DATE_FMT
        a.alignment = Alignment(horizontal="center", vertical="center")
        a.fill = GREEN_FILL
        a.border = THIN_BORDER
        a.protection = Protection(locked=True)
        a.font = Font(name="Calibri", size=10)

        for col in INPUT_COLS:
            cell = ws.cell(row, col, None)
            cell.protection = Protection(locked=False)
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=10)
            if col in (2, 5):
                cell.number_format = "0.00"
            if is_off:
                if col == 7:
                    cell.fill = GRAY_FILL
                elif col in (2, 3, 5):
                    cell.fill = GREEN_FILL
                else:
                    cell.fill = PatternFill()
            else:
                cell.fill = PatternFill()

        d_cell = ws.cell(row, 4, f"=B{row}-C{row}")
        d_cell.number_format = "0.00"
        d_cell.fill = GREEN_FILL
        d_cell.border = THIN_BORDER
        d_cell.protection = Protection(locked=True)
        d_cell.font = Font(name="Calibri", size=10)

        h = ws.cell(row, 8, f"=D{row}-E{row}")
        h.number_format = "0.00"
        h.border = THIN_BORDER
        h.protection = Protection(locked=True)
        h.font = Font(name="Calibri", size=10)

        prev_i = f"I{row - 1}" if row > DATA_START_ROW else "I13"
        i = ws.cell(row, 9, f"={prev_i}+H{row}-G{row}")
        i.number_format = MONEY_FMT
        i.border = THIN_BORDER
        i.protection = Protection(locked=True)
        i.font = Font(name="Calibri", size=10)

        for col in SELLER_COLS:
            cell = ws.cell(row, col, None)
            cell.protection = Protection(locked=False)
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=10)
            cell.number_format = "h:mm" if col in (11, 12) else "@"
            cell.fill = GREEN_FILL if is_off else PatternFill()

    def _clear_data_area(self, ws: Worksheet, start_row: int, end_row: int) -> None:
        for row in range(start_row, end_row + 1):
            for col in range(1, 16):
                cell = ws.cell(row, col)
                cell.value = None
                cell.fill = PatternFill()
                cell.protection = Protection(locked=True)

    def _clear_holiday_helper(self, ws: Worksheet) -> None:
        for row in range(1, 40):
            cell = ws.cell(row, HOLIDAY_LIST_COL)
            cell.value = None

    def _protect_sheet(self, ws: Worksheet) -> None:
        password = self.config.protection_password or None
        if password == "":
            password = None
        ws.protection.sheet = True
        ws.protection.enable()
        ws.protection.password = password
        ws.protection.selectUnlockedCells = True
        ws.protection.selectLockedCells = True
